"""Excel (.xlsx) import/export via openpyxl — optional, with a clear fallback.

If openpyxl is not installed, the loader/saver raise a descriptive
``RuntimeError`` telling the user how to enable it. This keeps the core engine
free of any hard third-party dependency (see docs/architecture.md).
"""

from __future__ import annotations

from pathlib import Path

from ..core.reference import index_to_col, to_a1  # noqa: F401
from ..core.sheet import Sheet
from ..core.workbook import Workbook

try:
    import openpyxl  # type: ignore

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - exercised only without the dep
    openpyxl = None
    HAS_OPENPYXL = False


_FALLBACK_MSG = (
    "Excel import/export requires 'openpyxl'. Install it with:\n"
    "    pip install openpyxl\n"
    "or install qcell's excel extra:  pip install qcell[excel]"
)


def load_xlsx(path: str | Path) -> Workbook:
    if not HAS_OPENPYXL:
        raise RuntimeError(_FALLBACK_MSG)
    path = Path(path)
    # data_only=False keeps formulas as text so qcell re-evaluates them itself.
    wb_x = openpyxl.load_workbook(path, data_only=False)
    wb = Workbook.__new__(Workbook)
    wb.sheets = []
    wb.active = 0
    for ws in wb_x.worksheets:
        sheet = Sheet(ws.title)
        # openpyxl returns formulas as "=..."; numbers/strings as-is.
        sheet.set_cells_bulk(
            (cell.row - 1, cell.column - 1, str(cell.value))
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None)
        wb.sheets.append(sheet)
    wb._add_default_if_empty()
    return wb


def save_xlsx(wb: Workbook, path: str | Path, *, values: bool = False) -> None:
    """Write a workbook to .xlsx.

    ``values=False`` (default) writes raw cell text, so formulas survive the
    round-trip into Excel. ``values=True`` writes computed values instead.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(_FALLBACK_MSG)
    path = Path(path)
    wb_x = openpyxl.Workbook()
    # Remove the default sheet openpyxl creates; we add our own.
    default = wb_x.active
    wb_x.remove(default)
    for sheet in wb.sheets:
        ws = wb_x.create_sheet(title=sheet.name[:31])  # Excel caps title at 31
        n_rows, n_cols = sheet.used_bounds()
        for r in range(n_rows):
            for c in range(n_cols):
                cell = sheet.get_cell(r, c)
                if cell is None:
                    continue
                if values:
                    val = sheet.get_value(r, c)
                    ws.cell(row=r + 1, column=c + 1, value=_coerce(val))
                else:
                    ws.cell(row=r + 1, column=c + 1, value=_excel_raw(cell))
    if not wb_x.worksheets:
        wb_x.create_sheet(title="Sheet1")
    wb_x.save(path)


def _excel_raw(cell) -> object:
    """Convert a qcell raw cell to a value openpyxl will write faithfully."""
    if cell.is_formula:
        return cell.raw  # already begins with '='
    return cell.literal()


def _coerce(val: object) -> object:
    from ..core.errors import CellError

    if isinstance(val, CellError):
        return str(val)
    return val
